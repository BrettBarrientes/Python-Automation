import boto3
import openpyxl

# Replace with your AWS credentials and region
aws_access_key_id = 'YOUR_AWS_ACCESS_KEY'
aws_secret_access_key = 'YOUR_AWS_SECRET_KEY'
region_name = 'YOUR_AWS_REGION'

# Create an S3 client
s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key, region_name=region_name)

# Open the Excel file
wb = openpyxl.load_workbook('path/to/your/excel/file.xlsx')
sheet = wb.active

# Iterate through each row in the Excel sheet
for i in range(1, sheet.max_row + 1):
    source_bucket = 'YOUR_SOURCE_BUCKET_NAME'
    destination_bucket = 'YOUR_DESTINATION_BUCKET_NAME'

    source_object_key = sheet.cell(row=i, column=1).value
    destination_object_key = sheet.cell(row=i, column=2).value

    try:
        # Copy the object from the source bucket to the destination bucket
        s3.copy_object(
            Bucket=destination_bucket,
            CopySource={'Bucket': source_bucket, 'Key': source_object_key},
            Key=destination_object_key
        )

        # Delete the original object from the source bucket
        s3.delete_object(Bucket=source_bucket, Key=source_object_key)

        print(f"Object {source_object_key} moved to {destination_bucket}/{destination_object_key} successfully.")
    except Exception as e:
        print(f"Error moving object: {e}")

print("All objects have been transferred and renamed successfully!")
