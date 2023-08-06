import os
import openpyxl
from openpyxl import load_workbook


# Script's purpose is to rename and move files from List 1 directory to List 2 directory.
# Instead of moving files one at a time, this script will move multiple files to different directories all at once in a loop.
# This can be used for transferring objects between S3 buckets as well

#Open the excel file

wb = openpyxl.load_workbook('*path set here to the excel file*') #Open the excel file
sheet = wb.active #Open the active sheet

List1 = [] #Create an empty list for List 1

List2 = [] #Create an empty list for List 2

for i in range(1, sheet.max_row + 1): 

    List1.append(sheet.cell(row=i, column=1).value) #i = index 0 C:/Users/Documents/test1.txt --> C:/Users/Documents/test_folder_01/Remaining useful life.txt
    List2.append(sheet.cell(row=i, column=2).value) #i = index 1 C:/Users/Documents/test2.txt --> C:/Users/Documents/test_folder_02/Remaining useful life.txt

for i in range(len(List1)): #len(List1) = 2
    os.rename(List1[i], List2[i]) #Rename the files in List 1 to List 2
    print(List1[i], List2[i]) #Print the files that were renamed

print("The files have been transferred and renamed successfully!")
